import openpyxl
import os.path as op

def save_file_modification(subject_id, total_blocks):
    #data_path = 'data path here' 
    subject_id_file_name = subject_id + '_emotional_faces_data.xlsx'
    subject_id_metafile_name = subject_id + '_emotional_faces_meta_data.xlsx'
    subject_id_training_name = subject_id + '_emotional_faces_training_data.xlsx'
    save_file = op.join(data_path, subject_id_file_name)
    save_metafile = op.join(data_path, subject_id_metafile_name)
    save_trainingfile = op.join(data_path, subject_id_training_name)

    data_wb = openpyxl.Workbook() # opening the excel file now 
    meta_data_wb = openpyxl.Workbook()
    training_data_wb = openpyxl.Workbook()

    ### Adding to main data workbook first ### 

    block_number_list = [str(i+1) for i in range(total_blocks)] # list of strings from 1 to #total blocks
    block_number_strings = [('block_' + block_number_list[i]) for i in range(len(block_number_list))] # list of blocks
    
    variable_strings = ['condition', 'cue_present', 'type_of_trial', 'stimulus_one', 'stimulus_two', 'systole_diastole', 'catch_trial', 'arrow_direction', 'correct_incorrect_late', 'iti']

    for ws_name in range(len(total_blocks)):
        data_wb.create_sheet(faces_strings[ws_name]) # creating a sheet to save each block 
        
        data_sheet = data_wb[faces_strings[ws_name]]
        
        for header in range(len(variable_strings)): # adding the headers
            data_sheet.cell(1, header+1, variable_strings[header])

    ### Adding to meta-data workbook ### 

    meta_variable_strings = ['subject_id', 'age', 'order', 'random_seed', 'session', 'colour_1', 'colour_2', 'colour_3'] # colours refer to the 
    meta_data_wb.create_sheet('meta_data')
    meta_data_sheet = data_wb['meta_data')

    for header in range(len(meta_variable_strings)):
        meta_data_sheet.cell(1, header+1, meta_variable_strings[header])

    ### Adding to training-data workbook ### 

    training_data_wb.create_sheet('training_data')
    training_data_sheet = training_data_wb['training_data']

    for header in range(len(variable_strings)): #pretty sure the same headers will be used, changed if there are any unique
        training_data_sheet.cell(1, header+1, variable_strings[header])
    
    return data_wb, meta_data_wb, training_data_wb, save_file, save_metafile, save_trainingfile
